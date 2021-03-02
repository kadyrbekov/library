from flask import Flask, render_template, request
from openpyxl import load_workbook
from sqlalchemy.orm.session import sessionmaker
from sqlalchemy.orm import scoped_session
from database import engine, book


app = Flask(__name__)

db = scoped_session(sessionmaker(bind=engine))


@app.route("/")
def homepage():
   return render_template("index.html")


@app.route("/books/")
def books():

    with engine.connect() as con:
        books = con.execute("""SELECT * FROM "Book"; """)
    return render_template("books.html", object_list=books)


@app.route("/authors")
def authors():
   
    with engine.connect() as con:
        authors = con.execute("""SELECT author FROM "Book";""")
    return render_template("authors.html", authors = authors)


@app.route("/book_add", methods=["POST", "GET"]) #GET?
def book_add():
    if request.method == "POST":
    #getting requests from fomr located at book_edit.html 
        id_num = request.form.get("id_num")
        name=request.form.get("name")
        author=request.form.get("author")
        image=request.form.get("image")

        db.execute(""" INSERT INTO "Book" (id_num, name, author, image) VALUES (:id_num, :name, :author, :image)""",
                {"id_num":id_num, "name": name, "author": author,"image": image}) 
        db.commit()

    return render_template("book_add.html")


@app.route("/book/<id>/")
def book(id):
    obj = db.execute(f"""SELECT * FROM "Book" WHERE id_num = {id}; """).first()
    db.commit()
    return render_template("book.html", obj=obj) # **kwargs


# @app.route("/<int:id>/", methods=["GET", "POST"])
# def db_book_update(id):
#     message = ''
#     if request.method == "POST":
#         name = request.form.get("tale")
#         author = request.form.get("author")
#         image = request.form.get("image")
#         db.execute(f'''
#             UPDATE "Book"
#             SET 
#                 name='{name}',
#                 author='{author}',
#                 image='{image}'
#             WHERE id={id};
#         ''')
#         db.commit()
#         message = "Изменения сохранены"

#     book_object = db.execute(f'SELECT * FROM "Book" WHERE id={id};').first()
#     return render_template("database_book_update.html", book_object=book_object, message=message)


# @app.route("/book/<num>/edit")
# def book_edit(num):
#     num = int(num) + 2
#     excel_file = load_workbook("tales.xlsx")
#     page = excel_file["Sheet1"]
#     tale = page[f"A{num}"]
#     author = page[f"B{num}"]
#     image = page[f"C{num}"]
#     obj = [tale.value, author.value, image.value, num]
#     return render_template("book_edit.html", obj=obj)


# @app.route("/book/<num>/save/", methods=["POST"])
# def book_save(num):
#     num = int(num)
#     excel_file = load_workbook("tales.xlsx")
#     page = excel_file["Sheet1"]
#     form = request.form
#     page[f"A{num}"] = form["tale"]
#     page[f"B{num}"] = form["author"]
#     page[f"C{num}"] = form["image"]
#     excel_file.save("tales.xlsx")
#     return "Сохранено!"

